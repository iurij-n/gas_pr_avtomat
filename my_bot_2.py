# import re
# from playwright.sync_api import Playwright, sync_playwright, expect
from playwright.sync_api import sync_playwright
import pygetwindow as gw
import pyautogui
import time
import openpyxl
from pathlib import Path

try:
    from creditor_data import CREDITOR_DATA
except ImportError:
    CREDITOR_DATA = dict()


DRAFT_TARGET_TYPE = "Заявление о вынесении судебного приказа (дубликата)"
DRAFT_TARGET_STATUS = "В процессе создания (черновик)"
GET_ELEMENT_TIMEOUT = 15000
PAUSE_BETWEEN_ATTEMPTS = 2
ALLOWED_EXTENSIONS = {".pdf",}
DEBTOR_DATA_FILE = 'Должники_20260318_121822_prepared.xlsx'


def close_dialog_if_exists():
    for _ in range(20):
        windows = gw.getWindowsWithTitle('Открытие')
        if windows:
            try:
                windows[0].activate()
            except Exception:
                pass
            else:
                pyautogui.press('esc')
                break
        time.sleep(0.5)


def get_element(page, identificator):
    el = page.locator(identificator)
    try:
        el.wait_for(state="visible", timeout=GET_ELEMENT_TIMEOUT)
    except:
        return None
    else:
        return el


def delete_first_matching_draft(page):

    try:
        page.goto("https://ej.sudrf.ru/")
    except:
        page.reload()

    page.get_by_title("Информация о движении поданных обращений в суд").click()
    page.get_by_role("button", name="Найти").click()

    try:
        table_container = page.locator(".table-responsive")
        table_container.wait_for(state="visible", timeout=15000)
    except:
        return

    rows = page.locator(".table-history tbody tr")
    rows_count = rows.count()

    for i in range(rows_count):
        row = rows.nth(i)
        row_type = row.locator("td").nth(3).inner_text().strip()
        row_status = row.locator("td").nth(4).inner_text().strip()

        if DRAFT_TARGET_TYPE in row_type and DRAFT_TARGET_STATUS in row_status:
            btn_delete = row.locator("td").nth(4).get_by_role("button", name="Удалить")
            if btn_delete.is_visible():
                btn_delete.click()
                confirm_btn = page.locator(".modal-content:has-text('Подтвердите удаление')").get_by_role("button", name="Удалить")
                if confirm_btn.is_visible(timeout=2000):
                    time.sleep(0.5)
                    confirm_btn.click()
                    time.sleep(0.5)
                return
        return


def fill_form(page, debtor_data):
    # print(f'\n\n{debtor_data}\n\n')
    try:
        page.goto("https://ej.sudrf.ru/")
    except:
        page.reload() # Если не вышло — просто обновляем страницу
    page.locator("#service-appeal").get_by_role("link", name="Подать обращение").click()
    page.get_by_role("link", name="Подать обращение").nth(2).click()
    page.get_by_role("link", name="Заявление о вынесении судебного приказа (дубликата)").click()
    # page.pause()
    # page.get_by_role("button", name="Я являюсь представителем").click()

    time.sleep(0.5)
    page.locator('button[name="Method"][value="2"]').click() # Кнопка "Я являюсь представителем"
    # page.locator("#Address_CourtNotices_Index").click()
    # page.locator("label").filter(has_text="Адрес для направления судебных повесток и иных судебных извещений:").click()
    # page.locator("#Address_CourtNotices_Index").click()
    page.locator("#Address_CourtNotices_Index").fill(CREDITOR_DATA.get('NOTIFY_POSTAL_CODE', ''))
    # page.get_by_role("textbox", name="Адрес для направления судебных повесток и иных судебных извещений:").click()
    # page.get_by_role("textbox", name="Адрес для направления судебных повесток и иных судебных извещений:").click()
    # page.get_by_role("textbox", name="Адрес для направления судебных повесток и иных судебных извещений:").press("ControlOrMeta+a")
    page.get_by_role("textbox", name="Адрес для направления судебных повесток и иных судебных извещений:").fill(CREDITOR_DATA.get('NOTIFY_ADDRESS', ''))
    # page.pause()
    page.get_by_role("button", name="Добавить файл").first.click()
    page.get_by_role("button", name="Выбрать файл").click()
    # page.get_by_role("button", name="Выбрать файл").set_input_files("cardxxxx.pdf")
    # page.get_by_role("button", name="Выбрать файл").set_input_files(r"C:\Users\triko\Downloads\Telegram Desktop\Desktop\Файлы_досье\Досье_8092417_Абеляшев Александр Анатольевич\cardxxxx.pdf")
    # 1. Запускаем ожидание события выбора файла
    with page.expect_file_chooser() as fc_info:
        # 2. Кликаем по вашей кнопке, которая "не является инпутом"
        # page.get_by_role("button", name="Выбрать файл").click()
        page.get_by_role("button", name="Выбрать файл").dispatch_event("click")

    # 3. Объект chooser сам найдет нужный скрытый инпут и вставит туда файл
    file_chooser = fc_info.value
    file_chooser.set_files(CREDITOR_DATA.get('AUTHORITY_CONFIRMATION_DOC', r''))
    page.get_by_role("button", name="Добавить").click()
    close_dialog_if_exists()
    
    # Кредитор
    page.get_by_role("button", name="Добавить заявителя").click()
    page.get_by_role("button", name="Юридическое лицо").click()
    # page.get_by_role("textbox", name="Наименование:").click()
    page.get_by_role("textbox", name="Наименование:").fill(CREDITOR_DATA.get('CREDITOR_NAME', ''))
    page.get_by_label("Процессуальный статус заявителя:").select_option("50710012")
    # page.get_by_role("textbox", name="ИНН:").click()
    page.get_by_role("textbox", name="ИНН:").fill(CREDITOR_DATA.get('CREDITOR_INN', ''))
    # page.get_by_role("textbox", name="ИНН:").click()
    # page.get_by_role("textbox", name="ОГРН:").click()
    page.get_by_role("textbox", name="ОГРН:").fill(CREDITOR_DATA.get('CREDITOR_OGRN', ''))
    # page.get_by_role("textbox", name="КПП:").click()
    page.get_by_role("textbox", name="КПП:").fill(CREDITOR_DATA.get('CREDITOR_KPP', ''))
    # page.locator("#Address_Legal_Index").click()
    page.locator("#Address_Legal_Index").fill(CREDITOR_DATA.get('LEGAL_POSTAL_CODE', ''))
    # page.locator("#Address_Legal_Index").press("Tab")
    page.get_by_role("textbox", name="Юридический адрес:").fill(CREDITOR_DATA.get('LEGAL_ADDRESS', ''))
    page.get_by_role("checkbox", name="Адрес фактического нахождения организации совпадает с юридическим адресом органи").check()
    # page.get_by_role("dialog", name="Данные заявителя").locator("#Email").click()
    page.get_by_role("dialog", name="Данные заявителя").locator("#Email").fill(CREDITOR_DATA.get('CREDITOR_EMAIL', ''))
    # page.get_by_role("dialog", name="Данные заявителя").locator("#Phone").click()
    page.get_by_role("dialog", name="Данные заявителя").locator("#Phone").fill(CREDITOR_DATA.get('CREDITOR_PHONE_NUMBER', ''))
    page.get_by_role("button", name="Сохранить").click()
    # page.pause()
    
    # Должник
    page.get_by_role("button", name="Добавить участника").click()
    page.get_by_role("button", name="Физическое лицо").click()
    # page.get_by_role("textbox", name="Фамилия").click()
    page.get_by_role("textbox", name="Фамилия").fill(debtor_data.get('Фамилия', ''))
    # page.get_by_role("textbox", name="Фамилия").press("Tab")
    page.get_by_role("textbox", name="Имя").fill(debtor_data.get('Имя', ''))
    # page.get_by_role("textbox", name="Имя").press("Tab")
    page.get_by_role("textbox", name="Отчество").fill(debtor_data.get('Отчество', ''))
    # page.get_by_role("textbox", name="Отчество").press("Tab")
    page.get_by_role("textbox", name="Введите дату рождения представляемого лица").fill(debtor_data.get('Дата рождения', ''))
    if debtor_data.get('Пол', '') == 'муж.':
        page.get_by_role("button", name="Мужской").click()
    if debtor_data.get('Пол', '') == 'жен.':
        page.get_by_role("button", name="Женский").click()
    # page.get_by_role("textbox", name="Введите адрес места рождения").click()
    page.get_by_role("textbox", name="Введите адрес места рождения").fill(debtor_data.get('МестоРождения', ''))
    # page.get_by_role("dialog", name="Данные участника процесса").locator("#Address_Permanent_Index").click()
    page.get_by_role("dialog", name="Данные участника процесса").locator("#Address_Permanent_Index").fill(debtor_data.get('РегистрацияИндекс', ''))
    # page.get_by_role("dialog", name="Данные участника процесса").locator("#Address_Permanent_Address").click()
    page.get_by_role("dialog", name="Данные участника процесса").locator("#Address_Permanent_Address").fill(debtor_data.get('РегистрацияАдрес', ''))
    # page.get_by_role("dialog", name="Данные участника процесса").locator("#Address_Actual_Index").click()
    page.get_by_role("dialog", name="Данные участника процесса").locator("#Address_Actual_Index").fill(debtor_data.get('МестоЖительстваИндекс', ''))
    # page.get_by_role("dialog", name="Данные участника процесса").locator("#Address_Actual_Address").click()
    page.get_by_role("dialog", name="Данные участника процесса").locator("#Address_Actual_Address").fill(debtor_data.get('МестоЖительстваАдрес', ''))
    page.get_by_role("dialog", name="Данные участника процесса").locator("#Type").select_option("1")
    # page.get_by_role("dialog", name="Данные участника процесса").locator("#Series").click()
    page.get_by_role("dialog", name="Данные участника процесса").locator("#Series").fill(debtor_data.get('ПаспортСерия', ''))
    # page.get_by_role("dialog", name="Данные участника процесса").locator("#Number").click()
    page.get_by_role("dialog", name="Данные участника процесса").locator("#Number").fill(debtor_data.get('ПаспортНомер', ''))
    # page.get_by_role("textbox", name="Дата выдачи").click()
    page.get_by_role("textbox", name="Дата выдачи").fill(debtor_data.get('ПаспортДатаВыдачи', ''))
    # page.get_by_role("dialog", name="Данные участника процесса").locator("#IssuedBy").click()
    page.get_by_role("dialog", name="Данные участника процесса").locator("#IssuedBy").fill(debtor_data.get('ПаспортКемВыдан', ''))
    # page.get_by_role("dialog", name="Данные участника процесса").locator("#IssuerId").click()
    page.get_by_role("dialog", name="Данные участника процесса").locator("#IssuerId").fill(debtor_data.get('ПаспортКодПодразделения', ''))
    page.get_by_role("checkbox", name="Полные данные участника процесса неизвестны").check()
    page.get_by_role("button", name="Сохранить").click()
    page.get_by_role("button", name="Выбрать суд").click()
    page.get_by_label("Регион:").select_option(debtor_data.get('РегионКод', ''))
    page.get_by_label("Судебный орган:").select_option(debtor_data.get('Код_Судебного_Участка', ''))
    page.get_by_role("button", name="Сохранить").click()
    # page.pause()
    # # page.get_by_role("link", name="Удалить файл").nth(2).click()
    page.get_by_role("button", name="Добавить файл").nth(1).click()
    page.get_by_role("button", name="Выбрать файл").click()
    # page.get_by_role("button", name="Выбрать файл").set_input_files("cardxxxx.pdf")
    # page.get_by_role("button", name="Выбрать файл").set_input_files(r"C:\Users\triko\Downloads\Telegram Desktop\Desktop\Файлы_досье\Досье_8092417_Абеляшев Александр Анатольевич\cardxxxx.pdf")
    # 1. Запускаем ожидание события выбора файла
    with page.expect_file_chooser() as fc_info:
        # 2. Кликаем по вашей кнопке, которая "не является инпутом"
        # page.get_by_role("button", name="Выбрать файл").click()
        page.get_by_role("button", name="Выбрать файл").dispatch_event("click")

    # 3. Объект chooser сам найдет нужный скрытый инпут и вставит туда файл
    file_chooser = fc_info.value
    file_chooser.set_files(debtor_data.get('Путь_Скан', ''))
    page.get_by_role("button", name="Добавить").click()
    close_dialog_if_exists()

    folder = Path(debtor_data.get('Путь_Досье', ''))

    if not folder.exists() or not folder.is_dir():
        print(f"Ошибка: Директория '{debtor_data.get('Путь_Досье', '')}' не найдена.")
        return

    for item in folder.iterdir():
        if item.is_file():
            ext = item.suffix.lower()
            if ext in ALLOWED_EXTENSIONS:
                page.get_by_role("button", name="Добавить файл").nth(1).click()
                page.get_by_role("button", name="Выбрать файл").click()
                with page.expect_file_chooser() as fc_info:
                    page.get_by_role("button", name="Выбрать файл").dispatch_event("click")
                file_chooser = fc_info.value
                file_chooser.set_files(item.resolve())
                page.get_by_role("button", name="Добавить").click()
                close_dialog_if_exists()
                time.sleep(0.5)

                # item.name - это имя файла с расширением (например, 'scan.pdf')
                # str(item.resolve()) - это абсолютный полный путь до файла
                # process_file(item.name, str(item.resolve()))
                # print(str(item.resolve()))
            else:
                print(f"Пропущен (недопустимый формат): {item.name}")

    #Уплата госпошлины
    page.get_by_role("checkbox", name="Квитанция об уплате государственной пошлины", exact=True).check()
    page.get_by_role("button", name="Добавить файл").nth(2).click()
    page.get_by_role("button", name="Выбрать файл").click()
    with page.expect_file_chooser() as fc_info:
        page.get_by_role("button", name="Выбрать файл").dispatch_event("click")
    file_chooser = fc_info.value
    file_chooser.set_files(debtor_data.get('Путь_Платежное_поручение', ''))
    page.get_by_role("button", name="Добавить").click()
    close_dialog_if_exists()
    time.sleep(0.5)
    
    #Согласия
    page.get_by_role("checkbox", name="Согласен на получение уведомлений на адрес электронной почты: iurij.").check()
    page.get_by_role("checkbox", name="Согласен на получение судебной корреспонденции на адрес, указанный для направлен").check()


    #Сформировать обращение
    page.get_by_role("button", name="Сформировать обращение").click()
    
    #Подать
    #page.get_by_role("button", name="Отправить").click()
    page.pause()
    
    # from playwright.sync_api import TimeoutError

# ... ваш код заполнения полей формы для текущего должника ...

    try:
        page.wait_for_selector('text="Ваше заявление успешно отправлено"', timeout=20000)

        # print(f"Успех: Заявление отправлено.")
        # Здесь записываем статус "Отправлено" в ваш Excel-файл
    except:
        pass
        # Попадаем сюда, если за 15 секунд заветный текст так и не появился на экране
        # print(f"Ошибка: Текст подтверждения не найден. Возможно, сбой на сайте.")
    else:
        return True
    return False
     


# def run(playwright: Playwright) -> None:
#     browser = playwright.chromium.launch(headless=False, slow_mo=1000)
#     context = browser.new_context(storage_state="auth.json")
#     page = context.new_page()
#     page.set_default_timeout(90000)
#     # #####################################################################################
    

#     # Открываем книгу и активный лист
#     file_path = 'Должники_20260318_121822_prepared.xlsx'
#     max_retries = 3
#     wb = openpyxl.load_workbook(file_path)
#     sheet = wb.active

#     # Читаем заголовки (первая строка), чтобы динамически собирать словарь
#     # Формат: {"Имя": 0, "Фамилия": 1, ...}
#     headers = {cell.value: idx for idx, cell in enumerate(sheet[1]) if cell.value}
    
#     # Проверяем, есть ли колонка "Статус". Если нет - добавляем.
#     if "Статус" not in headers:
#         status_col_idx = len(headers) + 1
#         sheet.cell(row=1, column=status_col_idx, value="Статус")
#         headers["Статус"] = status_col_idx - 1
#     else:
#         status_col_idx = headers["Статус"] + 1

#     # Идем по строкам, начиная со второй (пропуская заголовки)
#     for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        
#         # 1. Собираем словарь с данными должника
#         debtor_data = {
#             header: row[col_idx].value 
#             for header, col_idx in headers.items() 
#             if header != "Статус"
#         }

#         # Пропускаем строку, если заявление уже было успешно отправлено ранее
#         # (полезно, если скрипт упал и вы запустили его заново)
#         current_status = row[status_col_idx - 1].value
#         if current_status == "Отправлено":
#             continue

#         # 2. Валидация: проверяем, что все значения заполнены (нет None или пустых строк)
#         if not all(debtor_data.values()):
#             sheet.cell(row=row_idx, column=status_col_idx, value="Неполные данные")
#             wb.save(file_path) # Сохраняем промежуточный результат
#             print(f"Строка {row_idx}: Пропущена (неполные данные)")
#             continue

#         # 3. Цикл отправки с заданным количеством попыток
#         success = False
#         for attempt in range(1, max_retries + 1):
#             try:
#                 print(f"Строка {row_idx}, попытка {attempt}: Отправка {debtor_data.get('Фамилия')}...")
                
#                 # Вызываем функцию автоматизации
#                 # is_accepted = fill_and_submit_form(debtor_data)
#                 print(f'\n\n{debtor_data}\n\n')
#                 is_accepted = True
#                 if is_accepted:
#                     success = True
#                     break # Успех, выходим из цикла попыток
#                 else:
#                     print(f"Строка {row_idx}: Текст успешной отправки не найден.")
            
#             except Exception as e:
#                 print(f"Строка {row_idx}, попытка {attempt} завершилась ошибкой: {e}")
            
#             # Пауза перед следующей попыткой (чтобы страница/сервер успели "отвиснуть")
#             if attempt < max_retries:
#                 time.sleep(PAUSE_BETWEEN_ATTEMPTS)

#         # 4. Запись итогового статуса
#         if success:
#             sheet.cell(row=row_idx, column=status_col_idx, value="Отправлено")
#             print(f"Строка {row_idx}: УСПЕХ")
#         else:
#             sheet.cell(row=row_idx, column=status_col_idx, value="Не удалось отправить")
#             print(f"Строка {row_idx}: ПРОВАЛ после {max_retries} попыток")

#         # Сохраняем файл после каждой обработанной строки. 
#         # Это замедляет работу, но гарантирует, что при внезапном падении скрипта данные не потеряются.
#         wb.save(file_path)

#     print("Обработка файла завершена.")
#     # #####################################################################################
    
    
    
#     delete_first_matching_draft(page)
    
    
#     # fill_form(page, None)

#     # ---------------------
#     context.close()
#     browser.close()


if __name__ == '__main__':
    with sync_playwright() as playwright:
        # run(playwright)
        browser = playwright.chromium.launch(headless=False, slow_mo=300)
        context = browser.new_context(storage_state="auth.json")
        page = context.new_page()
        page.set_default_timeout(90000)
        # #####################################################################################
        

        # Открываем книгу и активный лист
        file_path = DEBTOR_DATA_FILE
        max_retries = 3
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Читаем заголовки (первая строка), чтобы динамически собирать словарь
        # Формат: {"Имя": 0, "Фамилия": 1, ...}
        headers = {cell.value: idx for idx, cell in enumerate(sheet[1]) if cell.value}
        
        # Проверяем, есть ли колонка "Статус". Если нет - добавляем.
        if "Статус" not in headers:
            status_col_idx = len(headers) + 1
            sheet.cell(row=1, column=status_col_idx, value="Статус")
            headers["Статус"] = status_col_idx - 1
        else:
            status_col_idx = headers["Статус"] + 1

        # Идем по строкам, начиная со второй (пропуская заголовки)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            
            # 1. Собираем словарь с данными должника
            debtor_data = {
                header: row[col_idx].value 
                for header, col_idx in headers.items() 
                if header != "Статус"
            }

            # Пропускаем строку, если заявление уже было успешно отправлено ранее
            # (полезно, если скрипт упал и вы запустили его заново)
            current_status = row[status_col_idx - 1].value
            if current_status == "Отправлено":
                continue

            # 2. Валидация: проверяем, что все значения заполнены (нет None или пустых строк)
            if not all(debtor_data.values()):
                sheet.cell(row=row_idx, column=status_col_idx, value="Неполные данные")
                wb.save(file_path) # Сохраняем промежуточный результат
                print(f"Строка {row_idx}: Пропущена (неполные данные)")
                continue

            # 3. Цикл отправки с заданным количеством попыток
            success = False
            for attempt in range(1, max_retries + 1):
                try:
                    print(f"Строка {row_idx}, попытка {attempt}: Отправка {debtor_data.get('Фамилия')}...")
                    
                    # Вызываем функцию автоматизации
                    delete_first_matching_draft(page)
                    is_accepted = fill_form(page, debtor_data)
                    
                    is_accepted = True
                    if is_accepted:
                        success = True
                        break # Успех, выходим из цикла попыток
                    else:
                        print(f"Строка {row_idx}: Текст успешной отправки не найден.")
                
                except Exception as e:
                    print(f"Строка {row_idx}, попытка {attempt} завершилась ошибкой: {e}")
                
                # Пауза перед следующей попыткой (чтобы страница/сервер успели "отвиснуть")
                if attempt < max_retries:
                    time.sleep(PAUSE_BETWEEN_ATTEMPTS)

            # 4. Запись итогового статуса
            if success:
                sheet.cell(row=row_idx, column=status_col_idx, value="Отправлено")
                print(f"Строка {row_idx}: УСПЕХ")
            else:
                sheet.cell(row=row_idx, column=status_col_idx, value="Не удалось отправить")
                print(f"Строка {row_idx}: ПРОВАЛ после {max_retries} попыток")

            # Сохраняем файл после каждой обработанной строки. 
            # Это замедляет работу, но гарантирует, что при внезапном падении скрипта данные не потеряются.
            wb.save(file_path)

        print("Обработка файла завершена.")
        # #####################################################################################
        
        
        
        
        
        
        # fill_form(page, None)

        # ---------------------
        page.close()
        context.close()
        browser.close()


