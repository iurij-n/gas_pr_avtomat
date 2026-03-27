from playwright.sync_api import sync_playwright
import pygetwindow as gw
import pyautogui
import time
import openpyxl
from pathlib import Path
import re


DRAFT_TARGET_TYPE = "Заявление о вынесении судебного приказа (дубликата)"  # Тип документа удаляемого черновика
DRAFT_TARGET_STATUS = "В процессе создания (черновик)"  # Статус удаляемого черновика
GET_ELEMENT_TIMEOUT = 30000  # Таймаут ожидания загрузки страницы, мс
PAUSE_BETWEEN_ATTEMPTS = 2  # Пауза между попытками составить заявление, сек
ALLOWED_EXTENSIONS = {".pdf", }  # Сет разрешенных расширений для прикрепляемых файлов
DEBTOR_DATA_FILE = ''  # Имя файла и данными для заполнения формы
SHARED_FILES_FOLDERS = ''  # Папка с общими файлами досье
AUTHORITY_CONFIRMATION_DOC = ''  # Доверенность
MAX_RETRIES = 3  # Количество попыток составить заявление
CREDITOR_DATA = {  # Данные кредитора
    'NOTIFY_POSTAL_CODE': '305000',
    'NOTIFY_ADDRESS': 'Курская область, г. Курск, ул. Марата, д. 21, этаж 2, помещение 1',
    'CREDITOR_NAME': 'ООО «ПКО «Центр Альтернативного Финансирования»',
    'CREDITOR_INN': '4632195224',
    'CREDITOR_OGRN': '1144632011380',
    'CREDITOR_KPP': '463201001',
    'LEGAL_POSTAL_CODE': '305000',
    'LEGAL_ADDRESS': 'Курская область, г. Курск, ул. Марата, д. 21, этаж 2, помещение 1',
    'CREDITOR_EMAIL': 'info@alternativacentr.ru',
    'CREDITOR_PHONE_NUMBER': '+79065730892',
}
USE_UCEP = True


def extract_application_data(page):
    page.locator("#ticket").wait_for(state="visible", timeout=10000)

    date_time = page.locator(
        "div.form-group:has(label:has-text('Дата и время отправки')) div[style]"
    ).inner_text()

    app_number = page.locator(
        "div.form-group:has(label:has-text('Номер')) div[style]"
    ).inner_text()

    return app_number, date_time


def close_dialog_if_exists():
    for _ in range(20):
        windows = gw.getWindowsWithTitle('Открытие')
        if windows:
            try:
                windows[0].activate()
            except Exception:
                pass
            else:
                pyautogui.press('esc')
                break
        time.sleep(0.5)


def attach_ucep(page, file_path):
    ucep_file_path = f'{file_path}.sig'
    if not Path(ucep_file_path).is_file():
        return

    with page.expect_file_chooser() as fc_info:
        page.get_by_role("button", name="Прикрепить файл УКЭП").dispatch_event("click")

    file_chooser = fc_info.value
    file_chooser.set_files(ucep_file_path)

    close_dialog_if_exists()


def get_element(page, identificator):
    el = page.locator(identificator)
    try:
        el.wait_for(state="visible", timeout=GET_ELEMENT_TIMEOUT)
    except:
        return None
    else:
        return el


def delete_first_matching_draft(page):

    try:
        page.goto("https://ej.sudrf.ru/")
    except:
        page.reload()

    page.get_by_title("Информация о движении поданных обращений в суд").click()
    page.get_by_role("button", name="Найти").click()

    try:
        table_container = page.locator(".table-responsive")
        table_container.wait_for(state="visible", timeout=GET_ELEMENT_TIMEOUT)
    except:
        return

    rows = page.locator(".table-history tbody tr")
    rows_count = rows.count()

    for i in range(rows_count):
        row = rows.nth(i)
        row_type = row.locator("td").nth(3).inner_text().strip()
        row_status = row.locator("td").nth(4).inner_text().strip()

        if DRAFT_TARGET_TYPE in row_type and DRAFT_TARGET_STATUS in row_status:
            btn_delete = row.locator("td").nth(4).get_by_role("button", name="Удалить")
            if btn_delete.is_visible():
                btn_delete.click()
                confirm_btn = page.locator(".modal-content:has-text('Подтвердите удаление')").get_by_role("button", name="Удалить")
                if confirm_btn.is_visible(timeout=GET_ELEMENT_TIMEOUT):
                    time.sleep(0.5)
                    confirm_btn.click()
                    time.sleep(0.5)
                return
        return


def fill_form(page, debtor_data):
    try:
        page.goto("https://ej.sudrf.ru/")
    except:
        page.reload()
    page.locator("#service-appeal").get_by_role("link", name="Подать обращение").click()
    page.get_by_role("link", name="Подать обращение").nth(2).click()
    page.get_by_role("link", name="Заявление о вынесении судебного приказа (дубликата)").click()
    time.sleep(0.5)
    page.locator('button[name="Method"][value="2"]').click()
    page.locator("#Address_CourtNotices_Index").fill(CREDITOR_DATA.get('NOTIFY_POSTAL_CODE', ''))
    page.get_by_role("textbox", name="Адрес для направления судебных повесток и иных судебных извещений:").fill(CREDITOR_DATA.get('NOTIFY_ADDRESS', ''))
    page.get_by_role("button", name="Добавить файл").first.click()
    page.get_by_role("button", name="Выбрать файл").click()

    with page.expect_file_chooser() as fc_info:
        page.get_by_role("button", name="Выбрать файл").dispatch_event("click")

    file_chooser = fc_info.value
    file_chooser.set_files(AUTHORITY_CONFIRMATION_DOC)
    close_dialog_if_exists()
    if USE_UCEP:
        attach_ucep(page, AUTHORITY_CONFIRMATION_DOC)
    page.get_by_role("button", name="Добавить").click()

    # Кредитор
    page.get_by_role("button", name="Добавить заявителя").click()
    page.get_by_role("button", name="Юридическое лицо").click()
    page.get_by_role("textbox", name="Наименование:").fill(CREDITOR_DATA.get('CREDITOR_NAME', ''))
    page.get_by_label("Процессуальный статус заявителя:").select_option("50710012")
    page.get_by_role("textbox", name="ИНН:").fill(CREDITOR_DATA.get('CREDITOR_INN', ''))
    page.get_by_role("textbox", name="ОГРН:").fill(CREDITOR_DATA.get('CREDITOR_OGRN', ''))
    page.get_by_role("textbox", name="КПП:").fill(CREDITOR_DATA.get('CREDITOR_KPP', ''))
    page.locator("#Address_Legal_Index").fill(CREDITOR_DATA.get('LEGAL_POSTAL_CODE', ''))
    page.get_by_role("textbox", name="Юридический адрес:").fill(CREDITOR_DATA.get('LEGAL_ADDRESS', ''))
    page.get_by_role("checkbox", name="Адрес фактического нахождения организации совпадает с юридическим адресом органи").check()
    page.get_by_role("dialog", name="Данные заявителя").locator("#Email").fill(CREDITOR_DATA.get('CREDITOR_EMAIL', ''))
    page.get_by_role("dialog", name="Данные заявителя").locator("#Phone").fill(CREDITOR_DATA.get('CREDITOR_PHONE_NUMBER', ''))
    page.get_by_role("button", name="Сохранить").click()

    # Должник
    page.get_by_role("button", name="Добавить участника").click()
    page.get_by_role("button", name="Физическое лицо").click()
    page.get_by_role("textbox", name="Фамилия").fill(debtor_data.get('Фамилия', ''))
    page.get_by_role("textbox", name="Имя").fill(debtor_data.get('Имя', ''))
    page.get_by_role("textbox", name="Отчество").fill(debtor_data.get('Отчество', ''))
    page.get_by_role("textbox", name="Введите дату рождения представляемого лица").fill(debtor_data.get('Дата рождения', ''))
    if debtor_data.get('Пол', '') == 'муж.':
        page.get_by_role("button", name="Мужской").click()
    if debtor_data.get('Пол', '') == 'жен.':
        page.get_by_role("button", name="Женский").click()
    page.get_by_role("textbox", name="Введите адрес места рождения").fill(debtor_data.get('МестоРождения', ''))
    page.get_by_role("dialog", name="Данные участника процесса").locator("#Address_Permanent_Index").fill(debtor_data.get('РегистрацияИндекс', ''))
    page.get_by_role("dialog", name="Данные участника процесса").locator("#Address_Permanent_Address").fill(debtor_data.get('РегистрацияАдрес', ''))
    page.get_by_role("dialog", name="Данные участника процесса").locator("#Address_Actual_Index").fill(debtor_data.get('МестоЖительстваИндекс', ''))
    page.get_by_role("dialog", name="Данные участника процесса").locator("#Address_Actual_Address").fill(debtor_data.get('МестоЖительстваАдрес', ''))
    page.get_by_role("dialog", name="Данные участника процесса").locator("#Type").select_option("1")
    page.get_by_role("dialog", name="Данные участника процесса").locator("#Series").fill(debtor_data.get('ПаспортСерия', ''))
    page.get_by_role("dialog", name="Данные участника процесса").locator("#Number").fill(debtor_data.get('ПаспортНомер', ''))
    page.get_by_role("textbox", name="Дата выдачи").fill(debtor_data.get('ПаспортДатаВыдачи', ''))
    page.get_by_role("dialog", name="Данные участника процесса").locator("#IssuedBy").fill(debtor_data.get('ПаспортКемВыдан', ''))
    page.get_by_role("dialog", name="Данные участника процесса").locator("#IssuerId").fill(debtor_data.get('ПаспортКодПодразделения', ''))
    page.get_by_role("checkbox", name="Полные данные участника процесса неизвестны").check()
    page.get_by_role("button", name="Сохранить").click()

    # Субебный участок
    page.get_by_role("button", name="Выбрать суд").click()
    page.get_by_label("Регион:").select_option(debtor_data.get('РегионКод', ''))
    page.get_by_label("Судебный орган:").select_option(debtor_data.get('Код_Судебного_Участка', ''))
    page.get_by_role("button", name="Сохранить").click()

    # Заявление
    page.get_by_role("button", name="Добавить файл").nth(1).click()
    page.get_by_role("button", name="Выбрать файл").click()

    with page.expect_file_chooser() as fc_info:
        page.get_by_role("button", name="Выбрать файл").dispatch_event("click")

    file_chooser = fc_info.value
    file_chooser.set_files(debtor_data.get('Путь_Скан', ''))
    close_dialog_if_exists()
    if USE_UCEP:
        attach_ucep(page, AUTHORITY_CONFIRMATION_DOC)
    page.get_by_role("button", name="Добавить").click()

    # Общие файлы досье
    folder = Path(SHARED_FILES_FOLDERS)

    if not folder.exists() or not folder.is_dir():
        print(f"Ошибка: Директория '{SHARED_FILES_FOLDERS}' не найдена.")
        return

    for item in folder.iterdir():
        if item.is_file():
            ext = item.suffix.lower()
            if ext in ALLOWED_EXTENSIONS:
                page.get_by_role("button", name="Добавить файл").nth(1).click()
                page.get_by_role("button", name="Выбрать файл").click()
                with page.expect_file_chooser() as fc_info:
                    page.get_by_role("button", name="Выбрать файл").dispatch_event("click")
                file_chooser = fc_info.value
                file_chooser.set_files(item.resolve())
                close_dialog_if_exists()
                if USE_UCEP:
                    attach_ucep(page, AUTHORITY_CONFIRMATION_DOC)
                page.get_by_role("button", name="Добавить").click()
                time.sleep(0.5)
            else:
                print(f"Пропущен (недопустимый формат): {item.name}")

    # Досье
    folder = Path(debtor_data.get('Путь_Досье', ''))

    if not folder.exists() or not folder.is_dir():
        print(f"Ошибка: Директория '{debtor_data.get('Путь_Досье', '')}' не найдена.")
        return

    for item in folder.iterdir():
        if item.is_file():
            ext = item.suffix.lower()
            if ext in ALLOWED_EXTENSIONS:
                page.get_by_role("button", name="Добавить файл").nth(1).click()
                page.get_by_role("button", name="Выбрать файл").click()
                with page.expect_file_chooser() as fc_info:
                    page.get_by_role("button", name="Выбрать файл").dispatch_event("click")
                file_chooser = fc_info.value
                file_chooser.set_files(item.resolve())
                close_dialog_if_exists()
                if USE_UCEP:
                    attach_ucep(page, AUTHORITY_CONFIRMATION_DOC)
                page.get_by_role("button", name="Добавить").click()
                time.sleep(0.5)
            else:
                print(f"Пропущен (недопустимый формат): {item.name}")

    # Уплата госпошлины
    page.get_by_role("checkbox", name="Квитанция об уплате государственной пошлины", exact=True).check()
    page.get_by_role("button", name="Добавить файл").nth(2).click()
    page.get_by_role("button", name="Выбрать файл").click()
    with page.expect_file_chooser() as fc_info:
        page.get_by_role("button", name="Выбрать файл").dispatch_event("click")
    file_chooser = fc_info.value
    file_chooser.set_files(debtor_data.get('Путь_Платежное_поручение', ''))
    close_dialog_if_exists()
    if USE_UCEP:
        attach_ucep(page, AUTHORITY_CONFIRMATION_DOC)
    page.get_by_role("button", name="Добавить").click()
    time.sleep(0.5)

    # Согласия
    page.get_by_role("checkbox", name=re.compile(r"^Согласен на получение уведомлений на адрес электронной почты")).check()
    page.get_by_role("checkbox", name="Согласен на получение судебной корреспонденции на адрес, указанный для направлен").check()

    # Сформировать обращение
    page.pause()
    # page.get_by_role("button", name="Сформировать обращение").click()

    # Подать
    # page.get_by_role("button", name="Отправить").click()

    try:
        page.wait_for_selector('text="Ваше заявление успешно отправлено"', timeout=GET_ELEMENT_TIMEOUT)
        # page.wait_for_selector('p:has-text("Ваше заявление успешно отправлено")', timeout=GET_ELEMENT_TIMEOUT)
    except:
        pass
    else:
        return True
    return False


if __name__ == '__main__':
    if not DEBTOR_DATA_FILE or not SHARED_FILES_FOLDERS or not AUTHORITY_CONFIRMATION_DOC:
        print('Ошибка. Значение констанот не указано.')

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False, slow_mo=300)
        context = browser.new_context(storage_state="auth.json")
        page = context.new_page()
        page.set_default_timeout(90000)

        file_path = DEBTOR_DATA_FILE
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        headers = {cell.value: idx for idx, cell in enumerate(sheet[1]) if cell.value}

        if "Статус" not in headers:
            status_col_idx = len(headers) + 1
            sheet.cell(row=1, column=status_col_idx, value="Статус")
            headers["Статус"] = status_col_idx - 1
        else:
            status_col_idx = headers["Статус"] + 1
            
        if "Дата заявки" not in headers:
            app_date_col_idx = len(headers) + 2
            sheet.cell(row=1, column=app_date_col_idx, value="Дата заявки")
            headers["Дата заявки"] = app_date_col_idx - 1
        else:
            app_date_col_idx = headers["Дата заявки"] + 1
            
        if "Номер заявки" not in headers:
            app_number_col_idx = len(headers) + 3
            sheet.cell(row=1, column=app_number_col_idx, value="Номер заявки")
            headers["Номер заявки"] = app_number_col_idx - 1
        else:
            app_number_col_idx = headers["Номер заявки"] + 1

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            debtor_data = {
                header: row[col_idx].value 
                for header, col_idx in headers.items() 
                if header != "Статус"
            }

            current_status = row[status_col_idx - 1].value
            if current_status == "Отправлено":
                continue

            if not all(debtor_data.values()):
                sheet.cell(row=row_idx, column=status_col_idx, value="Неполные данные")
                wb.save(file_path)
                print(f"Строка {row_idx}: Пропущена (неполные данные)")
                continue

            success = False
            for attempt in range(1, MAX_RETRIES + 1):
                try:
                    print(f"Строка {row_idx}, попытка {attempt}: Отправка {debtor_data.get('Фамилия')}...")
                    delete_first_matching_draft(page)
                    is_accepted = fill_form(page, debtor_data)

                    is_accepted = True
                    if is_accepted:
                        success = True
                        break
                    else:
                        print(f"Строка {row_idx}: Текст успешной отправки не найден.")

                except Exception as e:
                    print(f"Строка {row_idx}, попытка {attempt} завершилась ошибкой: {e}")

                if attempt < MAX_RETRIES:
                    time.sleep(PAUSE_BETWEEN_ATTEMPTS)

            if success:
                app_number, date_time = extract_application_data(page)
                sheet.cell(row=row_idx, column=status_col_idx, value="Отправлено")
                sheet.cell(row=row_idx, column=app_date_col_idx, value=date_time)
                sheet.cell(row=row_idx, column=app_number_col_idx, value=app_number)
                print(f"Строка {row_idx}: УСПЕХ")
            else:
                sheet.cell(row=row_idx, column=status_col_idx, value="Не удалось отправить")
                print(f"Строка {row_idx}: ПРОВАЛ после {MAX_RETRIES} попыток")
            wb.save(file_path)

        print("Обработка файла завершена.")
        page.close()
        context.close()
        browser.close()
